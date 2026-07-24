# -*- coding: utf-8 -*-
import json, re, unicodedata
import openpyxl

SRC = '../Description_des_champs.xlsx'

MEASURE_PREFIXES = ('I5F','Y5S','X7F','X9F','X8F','Y1F','A7S')

BASE_PAGES = {
    'RPE_DESCRIPTIONPUBLICS': ['Description des publics', 'Analyse par public'],
    'RPE_TAETPED': ['Accès / Présence en emploi', 'Analyse par public'],
    'RPE_SATISACO': ['Satisfaction DE', 'Analyse par public'],
    'RPE_FPRIO': ['Entrants en formation', 'Analyse par public'],
    'RPE_FTAE': ['Sortants de formation', 'Analyse par public'],
    'RPE_ENT': ['Taux de recours', 'Analyse par filière'],
    'RPE_TPODPO': ['Taux / Délai de pourvoi', 'Analyse par filière'],
}
BASE_SOURCES = {
    'RPE_DESCRIPTIONPUBLICS': ['Statistiques du Marché du Travail (STMT) - DARES / France Travail', 'Diagnostic Socio-Professionnel (DSP) - Opérateurs du RPE', 'Autres données France Travail'],
    'RPE_TAETPED': ['Statistiques du Marché du Travail (STMT) - DARES / France Travail', 'Déclaration Sociale Nominative (DSN) - GIP-MDS', 'Autres données France Travail'],
    'RPE_SATISACO': ['Enquête IPSOS', 'Autres données France Travail'],
    'RPE_FPRIO': ["Attestations d'Entrée en Stage (AES) - France Travail", 'Compte personnel de formation (CPF) - Caisse des Dépôts'],
    'RPE_FTAE': ["Attestations d'Entrée en Stage (AES) - France Travail", 'Compte personnel de formation (CPF) - Caisse des Dépôts', 'Déclaration Sociale Nominative (DSN) - GIP-MDS', 'Autres données France Travail'],
    'RPE_ENT': ['Déclaration Préalable à l\'Embauche (DPAE) - Urssaf / MSA', 'Déclaration Sociale Nominative (DSN) - GIP-MDS', "Offres d'emploi France Travail", 'Autres données France Travail'],
    'RPE_TPODPO': ["Offres d'emploi France Travail", 'Autres données France Travail'],
}

def clean(s):
    if s is None:
        return None
    s = str(s).replace('\xa0',' ').strip()
    s = re.sub(r'\s+\n', '\n', s)
    s = re.sub(r'[ \t]+', ' ', s)
    return s.strip()

def norm_key(s):
    s = clean(s) or ''
    s = s.lower()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s

def is_measure(tech):
    t = (tech or '').strip().upper()
    return t.startswith(MEASURE_PREFIXES)

def slugify(base_key, suffix=None):
    k = base_key
    if suffix:
        k = k + '-' + suffix
    return k

wb = openpyxl.load_workbook(SRC, data_only=True)

raw_entries = []  # list of dicts, one per kept row

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        tech = clean(row[1]) if len(row) > 1 else None
        libelle = clean(row[2]) if len(row) > 2 else None
        detail = clean(row[3]) if len(row) > 3 else None
        regle = clean(row[4]) if len(row) > 4 else None
        if not tech:
            continue
        # exclude technical-only fields
        blob = ' '.join([x or '' for x in [detail, regle]]).lower()
        if 'champ technique' in blob:
            continue
        if not libelle:
            continue
        raw_entries.append({
            'base': sheet, 'tech': tech, 'libelle': libelle,
            'detail': detail, 'regle': regle,
            'mesure': is_measure(tech),
        })

print('Total lignes retenues:', len(raw_entries))
mcount = sum(1 for e in raw_entries if e['mesure'])
print('Mesures:', mcount, 'Dimensions:', len(raw_entries)-mcount)

json.dump(raw_entries, open('raw_entries.json','w'), ensure_ascii=False, indent=1)
